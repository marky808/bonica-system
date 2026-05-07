#!/usr/bin/env python3
"""
請求書テンプレートに振込先情報を追加するスクリプト

使用方法:
    python scripts/add_bank_info_to_invoice.py <ファイルパス>

例:
    python scripts/add_bank_info_to_invoice.py ./templates/invoice_template.xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Border, Side
except ImportError:
    print("エラー: openpyxlがインストールされていません")
    print("インストール: pip install openpyxl")
    sys.exit(1)


def find_last_data_row(ws, start_row: int = 10) -> int:
    """
    A列の指定行以降を走査し、最終データ行を検出する

    Args:
        ws: ワークシート
        start_row: 走査開始行（デフォルト: 10）

    Returns:
        最終データ行の行番号
    """
    last_data_row = start_row - 1

    for row in range(start_row, ws.max_row + 2):  # max_row + 1まで確実にチェック
        cell_value = ws.cell(row=row, column=1).value  # A列
        if cell_value is None or str(cell_value).strip() == "":
            # 空セルが見つかった → 直前の行が最終データ行
            break
        last_data_row = row

    return last_data_row


def add_bank_info(file_path: str) -> None:
    """
    請求書テンプレートに振込先情報を追加する

    Args:
        file_path: Excelファイルのパス
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

    if not path.suffix.lower() in ['.xlsx', '.xlsm']:
        raise ValueError(f"サポートされていないファイル形式です: {path.suffix}")

    # ワークブックを読み込む（既存の書式を保持）
    wb = load_workbook(file_path)
    ws = wb.active

    # 最終データ行を検出
    last_data_row = find_last_data_row(ws, start_row=10)
    print(f"最終データ行: {last_data_row}行目")

    # 振込先情報の開始行（最終データ行 + 3）
    start_row = last_data_row + 3
    print(f"振込先情報開始行: {start_row}行目")

    # 振込先情報データ
    bank_info = [
        ("〈お振込先〉", None),           # N行目: タイトル
        ("金融機関名", "朝日信用金庫"),    # N+1行目
        ("支店名", "三郷支店"),            # N+2行目
        ("口座名義", "カ）ボニカ・アグリジェント"),  # N+3行目
        ("口座番号", "普通 0430910"),      # N+4行目
    ]

    # スタイル定義
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 振込先情報を書き込む
    for i, (label, value) in enumerate(bank_info):
        current_row = start_row + i

        # A列にラベルを書き込む
        cell_a = ws.cell(row=current_row, column=1, value=label)

        if i == 0:
            # タイトル行は太字
            cell_a.font = bold_font
        else:
            # データ行（N+1〜N+4）はB列に値を書き込み、罫線を付ける
            cell_b = ws.cell(row=current_row, column=2, value=value)

            # A列とB列に罫線を付ける
            cell_a.border = thin_border
            cell_b.border = thin_border

    # ファイルを保存
    wb.save(file_path)
    print(f"振込先情報を追加しました: {file_path}")
    print(f"  - 〈お振込先〉: {start_row}行目")
    print(f"  - 振込先詳細: {start_row + 1}〜{start_row + 4}行目")


def main():
    if len(sys.argv) < 2:
        print("使用方法: python add_bank_info_to_invoice.py <ファイルパス>")
        print("例: python add_bank_info_to_invoice.py ./templates/invoice.xlsx")
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        add_bank_info(file_path)
    except FileNotFoundError as e:
        print(f"エラー: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"エラー: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"予期しないエラー: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
